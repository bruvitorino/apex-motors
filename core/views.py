# Conteúdo FINAL CORRIGIDO para: core/views.py

# --- 1. IMPORTAÇÕES ---
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.db import models
from django.db.models import Max, Sum, Count
from django.contrib.auth.decorators import login_required
from django.db.models.functions import TruncMonth
from .models import Vendedor, Concessionaria, Cliente, Veiculo, Venda
import json
from datetime import datetime, timedelta

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from io import BytesIO


# ==================
# --- VENDEDORES ---
# ==================

@login_required
def lista_vendedores(request):
    vendedores = Vendedor.objects.all().order_by('id_vendedores')
    context = {'vendedores': vendedores}
    return render(request, 'core/lista_vendedores.html', context)


@login_required
def cria_vendedor(request):
    if request.method == 'POST':
        nome_vendedor = request.POST.get('nome')
        id_concessionaria = request.POST.get('id_concessionaria')
        max_id_result = Vendedor.objects.aggregate(max_id=Max('id_vendedores'))
        max_id = max_id_result['max_id'] if max_id_result['max_id'] is not None else 0
        next_id = max_id + 1
        now = timezone.now()
        Vendedor.objects.create(
            id_vendedores=next_id, nome=nome_vendedor, id_concessionarias_id=id_concessionaria,
            data_inclusao=now, data_atualizacao=now
        )
        return redirect('lista_vendedores')
    concessionarias = Concessionaria.objects.all().order_by('nome')
    context = {'concessionarias': concessionarias}
    return render(request, 'core/form_vendedor.html', context)


@login_required
def edita_vendedor(request, id):
    vendedor = Vendedor.objects.get(id_vendedores=id)
    if request.method == 'POST':
        vendedor.nome = request.POST.get('nome')
        vendedor.id_concessionarias_id = request.POST.get('id_concessionaria')
        vendedor.data_atualizacao = timezone.now()
        vendedor.save()
        return redirect('lista_vendedores')
    concessionarias = Concessionaria.objects.all().order_by('nome')
    context = {'vendedor': vendedor, 'concessionarias': concessionarias}
    return render(request, 'core/form_edita_vendedor.html', context)


@login_required
def exclui_vendedor(request, id):
    vendedor = Vendedor.objects.get(id_vendedores=id)
    if request.method == 'POST':
        vendedor.delete()
        return redirect('lista_vendedores')
    context = {'vendedor': vendedor}
    return render(request, 'core/confirma_exclusao.html', context)


# ==================
# --- CLIENTES ---
# ==================

@login_required
def lista_clientes(request):
    clientes = Cliente.objects.all().order_by('id_clientes')
    context = {'clientes': clientes}
    return render(request, 'core/lista_clientes.html', context)


@login_required
def cria_cliente(request):
    if request.method == 'POST':
        nome_cliente = request.POST.get('nome')
        endereco_cliente = request.POST.get('endereco')
        id_concessionaria = request.POST.get('id_concessionaria')
        max_id_result = Cliente.objects.aggregate(max_id=Max('id_clientes'))
        max_id = max_id_result['max_id'] if max_id_result['max_id'] is not None else 0
        next_id = max_id + 1
        now = timezone.now()
        Cliente.objects.create(
            id_clientes=next_id, nome=nome_cliente, endereco=endereco_cliente,
            id_concessionarias_id=id_concessionaria, data_inclusao=now, data_atualizacao=now
        )
        return redirect('lista_clientes')
   
    concessionarias = Concessionaria.objects.all().order_by('nome')
    context = {'concessionarias': concessionarias}
    return render(request, 'core/form_cliente.html', context)


@login_required
def edita_cliente(request, id):
    cliente = Cliente.objects.get(id_clientes=id)
   
    if request.method == 'POST':
        cliente.nome = request.POST.get('nome')
        cliente.endereco = request.POST.get('endereco')
        cliente.id_concessionarias_id = request.POST.get('id_concessionaria')
        cliente.data_atualizacao = timezone.now()
        cliente.save()
        return redirect('lista_clientes')

    concessionarias = Concessionaria.objects.all().order_by('nome')
    context = {
        'cliente': cliente,
        'concessionarias': concessionarias
    }
    return render(request, 'core/form_edita_cliente.html', context)


@login_required
def exclui_cliente(request, id):
    cliente = Cliente.objects.get(id_clientes=id)

    if request.method == 'POST':
        cliente.delete()
        return redirect('lista_clientes')

    context = {
        'cliente': cliente
    }
    return render(request, 'core/confirma_exclusao_cliente.html', context)


# ==================
# --- VENDAS ---
# ==================

@login_required
def lista_vendas(request):
    vendas = Venda.objects.select_related(
        'id_clientes', 'id_vendedores', 'id_veiculos'
    ).all().order_by('-id_vendas')

    context = {
        'vendas': vendas
    }
    return render(request, 'core/lista_vendas.html', context)


@login_required
def cria_venda(request):
    if request.method == 'POST':
        id_cliente = request.POST.get('id_cliente')
        id_veiculo = request.POST.get('id_veiculo')
        id_vendedor = request.POST.get('id_vendedor')
        id_concessionaria = request.POST.get('id_concessionaria')
        valor_pago = request.POST.get('valor_pago')
        data_venda = request.POST.get('data_venda')

        max_id_result = Venda.objects.aggregate(max_id=Max('id_vendas'))
        max_id = max_id_result['max_id'] if max_id_result['max_id'] is not None else 0
        next_id = max_id + 1

        now = timezone.now()

        Venda.objects.create(
            id_vendas=next_id,
            id_clientes_id=id_cliente,
            id_veiculos_id=id_veiculo,
            id_vendedores_id=id_vendedor,
            id_concessionarias_id=id_concessionaria,
            valor_pago=valor_pago,
            data_venda=data_venda,
            data_inclusao=now,
            data_atualizacao=now
        )

        return redirect('lista_vendas')

    clientes = Cliente.objects.all().order_by('nome')
    vendedores = Vendedor.objects.all().order_by('nome')
    veiculos = Veiculo.objects.all().order_by('nome')
    concessionarias = Concessionaria.objects.all().order_by('nome')
   
    context = {
        'clientes': clientes,
        'vendedores': vendedores,
        'veiculos': veiculos,
        'concessionarias': concessionarias
    }
    return render(request, 'core/form_venda.html', context)


@login_required
def edita_venda(request, id):
    venda = Venda.objects.get(id_vendas=id)
   
    if request.method == 'POST':
        venda.id_clientes_id = request.POST.get('id_cliente')
        venda.id_veiculos_id = request.POST.get('id_veiculo')
        venda.id_vendedores_id = request.POST.get('id_vendedor')
        venda.id_concessionarias_id = request.POST.get('id_concessionaria')
        venda.valor_pago = request.POST.get('valor_pago')
        venda.data_venda = request.POST.get('data_venda')
        venda.data_atualizacao = timezone.now()
        venda.save()
        return redirect('lista_vendas')

    clientes = Cliente.objects.all().order_by('nome')
    vendedores = Vendedor.objects.all().order_by('nome')
    veiculos = Veiculo.objects.all().order_by('nome')
    concessionarias = Concessionaria.objects.all().order_by('nome')
   
    context = {
        'venda': venda,
        'clientes': clientes,
        'vendedores': vendedores,
        'veiculos': veiculos,
        'concessionarias': concessionarias
    }
    return render(request, 'core/form_edita_venda.html', context)


@login_required
def exclui_venda(request, id):
    venda = Venda.objects.get(id_vendas=id)

    if request.method == 'POST':
        venda.delete()
        return redirect('lista_vendas')

    context = {
        'venda': venda
    }
    return render(request, 'core/confirma_exclusao_venda.html', context)


# ===============================================
# --- VIEWS PARA O SITE PÚBLICO E DASHBOARD ---
# ===============================================

def home(request):
    veiculos = Veiculo.objects.all().order_by('nome')
    context = {
        'veiculos': veiculos
    }
    return render(request, 'core/home.html', context)


# Dashboard de Gestão COM GRÁFICOS (apenas para admin)
@login_required
def dashboard(request):
    # Verificar se o usuário é admin/superuser
    if not request.user.is_superuser:
        return redirect('home')
   
    # Estatísticas gerais
    total_vendas = Venda.objects.count()
    total_clientes = Cliente.objects.count()
    total_vendedores = Vendedor.objects.count()
    faturamento_total = Venda.objects.aggregate(total=Sum('valor_pago'))['total'] or 0
   
    # Ranking de vendedores (top 5)
    ranking_vendedores = Venda.objects.values('id_vendedores__nome').annotate(
        total_vendas=Count('id_vendas'),
        total_faturamento=Sum('valor_pago')
    ).order_by('-total_vendas')[:5]
   
    # Veículos mais vendidos (top 5)
    veiculos_mais_vendidos = Venda.objects.values('id_veiculos__nome', 'id_veiculos__tipo').annotate(
        quantidade=Count('id_vendas'),
        faturamento=Sum('valor_pago')
    ).order_by('-quantidade')[:5]
   
    # Performance por concessionária
    performance_concessionarias = Venda.objects.values('id_vendedores__id_concessionarias__nome').annotate(
        total_vendas=Count('id_vendas'),
        faturamento=Sum('valor_pago')
    ).order_by('-total_vendas')[:10]
   
    # Vendas por mês (últimos 6 meses)
    seis_meses_atras = datetime.now() - timedelta(days=180)
    vendas_por_mes = Venda.objects.filter(data_venda__gte=seis_meses_atras).annotate(
        mes=TruncMonth('data_venda')
    ).values('mes').annotate(
        quantidade=Count('id_vendas'),
        faturamento=Sum('valor_pago')
    ).order_by('mes')

    # ===================================
    # PREPARAR DADOS PARA OS GRÁFICOS
    # ===================================
    
    # Gráfico de Vendas Mensais
    meses_labels = []
    vendas_count = []
    receita_values = []
    
    for venda in vendas_por_mes:
        meses_labels.append(venda['mes'].strftime('%B %Y'))
        vendas_count.append(venda['quantidade'])
        receita_values.append(float(venda['faturamento']))
    
    # Gráfico de Veículos
    veiculos_labels = []
    veiculos_data = []
    veiculos_colors = [
        '#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', 
        '#9966FF', '#FF9F40', '#FF6384'
    ]
    
    for veiculo in veiculos_mais_vendidos:
        veiculos_labels.append(veiculo['id_veiculos__nome'])
        veiculos_data.append(veiculo['quantidade'])
    
    # Gráfico de Concessionárias
    concessionarias_labels = []
    concessionarias_data = []
    
    for conc in performance_concessionarias:
        nome_conc = conc['id_vendedores__id_concessionarias__nome'] or 'Sem nome'
        concessionarias_labels.append(nome_conc[:20])
        concessionarias_data.append(conc['total_vendas'])
    
    # Gráfico de Top Vendedores
    vendedores_labels = []
    vendedores_data = []
    
    for vendedor in ranking_vendedores:
        vendedores_labels.append(vendedor['id_vendedores__nome'])
        vendedores_data.append(vendedor['total_vendas'])
   
    context = {
        'total_vendas': total_vendas,
        'total_clientes': total_clientes,
        'total_vendedores': total_vendedores,
        'faturamento_total': faturamento_total,
        'ranking_vendedores': ranking_vendedores,
        'veiculos_mais_vendidos': veiculos_mais_vendidos,
        'performance_concessionarias': performance_concessionarias,
        'vendas_por_mes': vendas_por_mes,
        # Dados para Chart.js (em JSON)
        'meses_labels': json.dumps(meses_labels),
        'vendas_count': json.dumps(vendas_count),
        'receita_values': json.dumps(receita_values),
        'veiculos_labels': json.dumps(veiculos_labels),
        'veiculos_data': json.dumps(veiculos_data),
        'veiculos_colors': json.dumps(veiculos_colors),
        'concessionarias_labels': json.dumps(concessionarias_labels),
        'concessionarias_data': json.dumps(concessionarias_data),
        'vendedores_labels': json.dumps(vendedores_labels),
        'vendedores_data': json.dumps(vendedores_data),
    }
   
    return render(request, 'core/dashboard.html', context)

@login_required
def exportar_excel(request):
    """
    Exporta os dados de vendas para Excel
    """
    # Verificar se é admin
    if not request.user.is_superuser:
        return redirect('home')
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendas Apex Motors"
    
    # Estilo do cabeçalho
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Cabeçalhos
    headers = [
        'ID Venda', 'Data Venda', 'Veículo', 'Cliente', 
        'Vendedor', 'Concessionária', 'Valor Pago'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Buscar dados
    vendas = Venda.objects.select_related(
        'id_veiculos', 'id_clientes', 'id_vendedores', 
        'id_vendedores__id_concessionarias'
    ).all().order_by('-data_venda')
    
    # Preencher dados
    for row_num, venda in enumerate(vendas, 2):
        ws.cell(row=row_num, column=1, value=venda.id_vendas)
        ws.cell(row=row_num, column=2, value=venda.data_venda.strftime('%d/%m/%Y'))
        ws.cell(row=row_num, column=3, value=venda.id_veiculos.nome if venda.id_veiculos else '')
        ws.cell(row=row_num, column=4, value=venda.id_clientes.nome if venda.id_clientes else '')
        ws.cell(row=row_num, column=5, value=venda.id_vendedores.nome if venda.id_vendedores else '')
        ws.cell(row=row_num, column=6, value=venda.id_vendedores.id_concessionarias.nome if venda.id_vendedores and venda.id_vendedores.id_concessionarias else '')
        ws.cell(row=row_num, column=7, value=float(venda.valor_pago))
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 18
    
    # Formatar coluna de valores como moeda
    for row_num in range(2, len(vendas) + 2):
        ws.cell(row=row_num, column=7).number_format = 'R$ #,##0.00'
    
    # Preparar resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=vendas_apex_motors_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    wb.save(response)
    return response


@login_required
def exportar_pdf(request):
    """
    Exporta relatório da dashboard em PDF
    """
    # Verificar se é admin
    if not request.user.is_superuser:
        return redirect('home')
    
    # Criar buffer
    buffer = BytesIO()
    
    # Criar documento PDF
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#0066CC'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#333333'),
        spaceAfter=12,
        spaceBefore=12
    )
    
    # Título
    title = Paragraph("Relatório de Vendas<br/>Apex Motors", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    # Data do relatório
    data_relatorio = Paragraph(
        f"<b>Data do Relatório:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        styles['Normal']
    )
    elements.append(data_relatorio)
    elements.append(Spacer(1, 0.3*inch))
    
    # Estatísticas gerais
    total_vendas = Venda.objects.count()
    total_clientes = Cliente.objects.count()
    total_vendedores = Vendedor.objects.count()
    faturamento_total = Venda.objects.aggregate(total=Sum('valor_pago'))['total'] or 0
    
    # Tabela de estatísticas
    elements.append(Paragraph("Estatísticas Gerais", heading_style))
    
    stats_data = [
        ['Métrica', 'Valor'],
        ['Total de Vendas', f'{total_vendas:,}'],
        ['Total de Clientes', f'{total_clientes:,}'],
        ['Total de Vendedores', f'{total_vendedores:,}'],
        ['Faturamento Total', f'R$ {faturamento_total:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')]
    ]
    
    stats_table = Table(stats_data, colWidths=[3*inch, 3*inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066CC')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(stats_table)
    elements.append(Spacer(1, 0.5*inch))
    
    # Top 5 Vendedores
    elements.append(Paragraph("Top 5 Vendedores", heading_style))
    
    ranking_vendedores = Venda.objects.values('id_vendedores__nome').annotate(
        total_vendas=Count('id_vendas'),
        total_faturamento=Sum('valor_pago')
    ).order_by('-total_vendas')[:5]
    
    vendedores_data = [['Vendedor', 'Total de Vendas', 'Faturamento']]
    for v in ranking_vendedores:
        vendedores_data.append([
            v['id_vendedores__nome'],
            str(v['total_vendas']),
            f"R$ {v['total_faturamento']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        ])
    
    vendedores_table = Table(vendedores_data, colWidths=[2.5*inch, 1.5*inch, 2*inch])
    vendedores_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066CC')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(vendedores_table)
    elements.append(Spacer(1, 0.5*inch))
    
    # Top 5 Veículos
    elements.append(Paragraph("Top 5 Veículos Mais Vendidos", heading_style))
    
    veiculos_mais_vendidos = Venda.objects.values('id_veiculos__nome').annotate(
        quantidade=Count('id_vendas'),
        faturamento=Sum('valor_pago')
    ).order_by('-quantidade')[:5]
    
    veiculos_data = [['Veículo', 'Quantidade', 'Faturamento']]
    for v in veiculos_mais_vendidos:
        veiculos_data.append([
            v['id_veiculos__nome'],
            str(v['quantidade']),
            f"R$ {v['faturamento']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        ])
    
    veiculos_table = Table(veiculos_data, colWidths=[2.5*inch, 1.5*inch, 2*inch])
    veiculos_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066CC')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(veiculos_table)
    
    # Construir PDF
    doc.build(elements)
    
    # Preparar resposta
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=relatorio_apex_motors_{datetime.now().strftime("%Y%m%d")}.pdf'
    
    return response

